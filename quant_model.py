#!/usr/bin/env python3
"""
Deterministic quant layer for the AI supply-chain knowledge base.

The model ranks graph tickers with three transparent factors:
  * structural bottleneck score from bottleneck_analysis.py
  * trailing momentum over --mom-days
  * relative strength from a multi-window cross-sectional return rank

It can print today's rule-derived levels or run a walk-forward backtest with cash reserve,
slippage, layer exposure limits, optional stop-loss simulation, and optional portfolio circuit
breaker. All prices and levels are formulas, not predictions.

Idea-generation only. Not investment advice. You decide and place every order.
"""
import argparse
import csv
import math
import os
import sys
from collections import defaultdict
from copy import copy

HERE = os.path.dirname(os.path.abspath(__file__))
HIST = os.path.join(HERE, "_tools", "prices_history.csv")
XLSX = os.path.join(HERE, "AI-Supply-Chain-Network.xlsx")
TRADING_DAYS = 252.0


# ---------------- data loading ----------------
def load_history(path=HIST):
    """Return {ticker: [(date, close), ...sorted]} from the long-format CSV."""
    series = defaultdict(list)
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            try:
                series[row["ticker"].upper()].append((row["date"], float(row["close"])))
            except (ValueError, KeyError, TypeError):
                continue
    for ticker in series:
        series[ticker].sort()
    return dict(series)


def load_bottleneck():
    """Return {ticker: structural_score} from the live graph engine."""
    try:
        sys.path.insert(0, HERE)
        import bottleneck_analysis as ba
        from openpyxl import load_workbook

        nodes, edges = ba.load()
        rows = ba.analyze(nodes, edges)
        wb = load_workbook(ba.XLSX, data_only=True, read_only=True)
        id2tic = {
            r[0]: str(r[3] or "").upper().strip()
            for r in wb["Nodes"].iter_rows(min_row=2, values_only=True)
            if r and r[0]
        }
        out = {}
        for row in rows:
            ticker = id2tic.get(row["id"], "")
            if ticker and ticker != "PRIVATE":
                out[ticker] = row["score"]
        return out
    except Exception as exc:
        print(f"  (bottleneck scores unavailable: {exc}; structural factor = 0)", file=sys.stderr)
        return {}


def load_metadata():
    """Return ticker metadata used for exposure caps and reporting."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(XLSX, data_only=True, read_only=True)
        out = {}
        for r in wb["Nodes"].iter_rows(min_row=2, values_only=True):
            if not r or not r[0] or r[2] != "Company":
                continue
            ticker = str(r[3] or "").upper().strip()
            if not ticker or ticker == "PRIVATE":
                continue
            out[ticker] = {
                "node_id": r[0],
                "name": r[1] or ticker,
                "layer": r[4],
                "layer_name": r[5] or "Unknown",
                "segment": r[6] or "",
                "country": r[7] or "",
            }
        return out
    except Exception:
        return {}


def parse_excludes(raw):
    return {t.strip().upper() for t in (raw or "").split(",") if t.strip()}


# ---------------- math helpers ----------------
def zscore(values):
    """dict -> dict of z-scores. Empty/flat input -> zeros."""
    vals = list(values.values())
    if not vals:
        return {}
    mean = sum(vals) / len(vals)
    var = sum((v - mean) ** 2 for v in vals) / len(vals)
    sd = math.sqrt(var)
    if sd == 0:
        return {k: 0.0 for k in values}
    return {k: (v - mean) / sd for k, v in values.items()}


def percentile_scores(values):
    """Map values to roughly [-1, 1] cross-sectional percentile scores."""
    if not values:
        return {}
    ordered = sorted(values.items(), key=lambda kv: kv[1])
    n = len(ordered)
    if n == 1:
        return {ordered[0][0]: 0.0}
    return {k: (2.0 * i / (n - 1)) - 1.0 for i, (k, _) in enumerate(ordered)}


def momentum(closes, lookback):
    """Trailing simple return over lookback bars. Needs lookback+1 points."""
    if len(closes) < lookback + 1:
        return None
    start, end = closes[-lookback - 1], closes[-1]
    return (end / start - 1.0) if start else None


def sma(closes, n):
    if len(closes) < n:
        return None
    return sum(closes[-n:]) / n


def levels(price, stop, target):
    """Rule-derived levels from user fractions. Not a forecast."""
    return {
        "entry_limit": round(price, 2),
        "stop": round(price * (1 - stop), 2),
        "target": round(price * (1 + target), 2),
        "r_multiple": round(target / stop, 2) if stop else None,
    }


def aligned_dates(hist):
    return sorted({d for series in hist.values() for d, _ in series})


def price_on(closes_by_ticker, dates, ticker, date_index):
    """Nearest prior close on or before dates[date_index]."""
    for j in range(date_index, -1, -1):
        price = closes_by_ticker.get(ticker, {}).get(dates[j])
        if price:
            return price
    return None


def closes_until(closes_by_ticker, dates, ticker, date_index):
    return [closes_by_ticker[ticker][d] for d in dates[: date_index + 1] if d in closes_by_ticker[ticker]]


def factor_snapshot(hist, struct, dates, date_index, mom_days, universe=None):
    """Compute factors using only data available through date_index."""
    closes_by_ticker = {t: dict(series) for t, series in hist.items()}
    universe = set(universe or hist.keys())
    mom = {}
    multi_window_returns = defaultdict(list)

    for ticker in sorted(universe):
        if ticker not in hist:
            continue
        closes = closes_until(closes_by_ticker, dates, ticker, date_index)
        mv = momentum(closes, mom_days)
        if mv is not None:
            mom[ticker] = mv
        for window in (21, 63, 126):
            rv = momentum(closes, window)
            if rv is not None:
                multi_window_returns[ticker].append(rv)

    raw_rs = {
        ticker: sum(vals) / len(vals)
        for ticker, vals in multi_window_returns.items()
        if vals and ticker in mom
    }
    rs = percentile_scores(raw_rs)
    struct_subset = {ticker: struct.get(ticker, 0.0) for ticker in mom}
    return struct_subset, mom, rs


def composite(struct, mom, rel_strength, w_struct, w_mom, w_rs):
    """Combine distinct factor dictionaries into one composite score."""
    zs = zscore(struct)
    zm = zscore(mom)
    keys = set(struct) | set(mom) | set(rel_strength)
    return {
        ticker: w_struct * zs.get(ticker, 0.0)
        + w_mom * zm.get(ticker, 0.0)
        + w_rs * rel_strength.get(ticker, 0.0)
        for ticker in keys
    }


def select_portfolio(ranked, metadata, top, gross_exposure, max_layer_weight):
    """Pick top names while respecting a max weight per graph layer."""
    if top <= 0 or gross_exposure <= 0:
        return {}, {}
    target_weight = gross_exposure / top
    weights = {}
    layer_weights = defaultdict(float)

    for ticker in ranked:
        layer = metadata.get(ticker, {}).get("layer_name", "Unknown")
        if layer_weights[layer] + target_weight > max_layer_weight + 1e-12:
            continue
        weights[ticker] = target_weight
        layer_weights[layer] += target_weight
        if len(weights) >= top:
            break
    return weights, dict(layer_weights)


# ---------------- signals ----------------
def run_signals(args):
    hist = load_history()
    struct = load_bottleneck()
    metadata = load_metadata()
    universe = (set(metadata) & set(hist)) - parse_excludes(args.exclude)
    dates = aligned_dates(hist)
    closes_by_ticker = {t: dict(series) for t, series in hist.items()}
    struct_f, mom_f, rs_f = factor_snapshot(hist, struct, dates, len(dates) - 1, args.mom_days, universe)
    comp = composite(struct_f, mom_f, rs_f, args.w_struct, args.w_mom, args.w_rs)
    ranked = sorted(comp.items(), key=lambda kv: -kv[1])

    print("=" * 108)
    print(
        f"COMPOSITE SIGNALS  struct x{args.w_struct} + mom x{args.w_mom} + relstr x{args.w_rs}; "
        f"mom={args.mom_days}d"
    )
    print("Relative strength is a multi-window cross-sectional rank. Levels are arithmetic rules, not forecasts.")
    print("=" * 108)
    print(
        f"{'#':<3}{'Ticker':<8}{'Composite':>10}{'Struct':>8}{'Mom%':>9}{'RS':>7}  "
        f"{'Layer':<24}{'Price':>11}{'Entry<=':>10}{'Stop':>9}{'Target':>9}"
    )
    print("-" * 108)
    for i, (ticker, score) in enumerate(ranked[: args.top if args.top else len(ranked)], 1):
        price = price_on(closes_by_ticker, dates, ticker, len(dates) - 1)
        lv = levels(price, args.stop, args.target) if price else {}
        layer = metadata.get(ticker, {}).get("layer_name", "")[:22]
        print(
            f"{i:<3}{ticker:<8}{score:>10.2f}{struct_f.get(ticker, 0):>8.1f}"
            f"{mom_f.get(ticker, 0) * 100:>8.1f}%{rs_f.get(ticker, 0):>7.2f}"
            f"  {layer:<22}{('$' + format(price, '.2f')) if price else '-':>11}"
            f"{('$' + format(lv.get('entry_limit'), '.2f')) if price else '-':>10}"
            f"{('$' + format(lv.get('stop'), '.2f')) if price else '-':>9}"
            f"{('$' + format(lv.get('target'), '.2f')) if price else '-':>9}"
        )
    print("\nIdea-generation only. Not investment advice. You decide and place every order.")


# ---------------- backtest ----------------
def returns_from_curve(curve):
    out = []
    for i in range(1, len(curve)):
        prev, cur = curve[i - 1][1], curve[i][1]
        if prev:
            out.append(cur / prev - 1.0)
    return out


def cagr(curve):
    if len(curve) < 2 or curve[-1][1] <= 0:
        return 0.0
    years = len(curve) / TRADING_DAYS
    return curve[-1][1] ** (1 / years) - 1


def maxdd(curve):
    peak = 0.0
    worst = 0.0
    for _, value in curve:
        peak = max(peak, value)
        if peak:
            worst = min(worst, value / peak - 1)
    return worst


def sharpe(curve):
    rets = returns_from_curve(curve)
    if len(rets) < 2:
        return 0.0
    mean = sum(rets) / len(rets)
    var = sum((r - mean) ** 2 for r in rets) / (len(rets) - 1)
    sd = math.sqrt(var)
    return (mean / sd) * math.sqrt(TRADING_DAYS) if sd else 0.0


def run_backtest(args, quiet=False):
    hist = load_history()
    struct = load_bottleneck()
    metadata = load_metadata()
    universe = (set(metadata) & set(hist)) - parse_excludes(args.exclude)
    dates = aligned_dates(hist)
    closes_by_ticker = {t: dict(series) for t, series in hist.items()}
    tickers = [t for t in universe if len(hist.get(t, [])) > args.mom_days + 5]

    if len(dates) < args.mom_days + args.rebal + 5 or len(tickers) < 2:
        if not quiet:
            print("Not enough history. Run _tools/fetch_history.py first.")
        return None

    gross = max(0.0, min(1.0, 1.0 - args.cash_reserve))
    max_layer = max(0.0, min(1.0, args.max_layer_weight))
    slippage = max(0.0, args.slippage_bps) / 10000.0
    portfolio_value = 1.0
    benchmark_value = 1.0
    equity = []
    benchmark_curve = []
    weights = {}
    entry_prices = {}
    holdings = []
    turnover_total = 0.0
    stop_count = 0
    circuit_count = 0
    paused_until_rebalance = False
    peak = 1.0

    for di in range(args.mom_days, len(dates)):
        day = dates[di]

        if weights:
            daily_return = 0.0
            stopped = []
            for ticker, weight in list(weights.items()):
                p0 = price_on(closes_by_ticker, dates, ticker, di - 1)
                p1 = price_on(closes_by_ticker, dates, ticker, di)
                if not p0 or not p1:
                    continue
                daily_return += weight * (p1 / p0 - 1.0)
                if args.backtest_stop and entry_prices.get(ticker) and p1 <= entry_prices[ticker] * (1 - args.backtest_stop):
                    stopped.append(ticker)
            portfolio_value *= 1.0 + daily_return
            if stopped:
                exit_cost = slippage * sum(weights[t] for t in stopped)
                portfolio_value *= 1.0 - exit_cost
                for ticker in stopped:
                    weights.pop(ticker, None)
                    entry_prices.pop(ticker, None)
                stop_count += len(stopped)

        if args.benchmark and args.benchmark in closes_by_ticker:
            bp0 = price_on(closes_by_ticker, dates, args.benchmark, di - 1)
            bp1 = price_on(closes_by_ticker, dates, args.benchmark, di)
            if bp0 and bp1:
                benchmark_value *= 1.0 + (bp1 / bp0 - 1.0)
        else:
            rs = []
            for ticker in tickers:
                p0 = price_on(closes_by_ticker, dates, ticker, di - 1)
                p1 = price_on(closes_by_ticker, dates, ticker, di)
                if p0 and p1:
                    rs.append(p1 / p0 - 1.0)
            if rs:
                benchmark_value *= 1.0 + sum(rs) / len(rs)

        peak = max(peak, portfolio_value)
        if args.portfolio_stop and weights and portfolio_value / peak - 1.0 <= -args.portfolio_stop:
            portfolio_value *= 1.0 - slippage * sum(weights.values())
            weights = {}
            entry_prices = {}
            paused_until_rebalance = True
            circuit_count += 1
            peak = portfolio_value

        equity.append((day, portfolio_value))
        benchmark_curve.append((day, benchmark_value))

        if (di - args.mom_days) % args.rebal == 0:
            struct_f, mom_f, rs_f = factor_snapshot(hist, struct, dates, di, args.mom_days, tickers)
            comp = composite(struct_f, mom_f, rs_f, args.w_struct, args.w_mom, args.w_rs)
            ranked = [ticker for ticker, _ in sorted(comp.items(), key=lambda kv: -kv[1])]
            new_weights, layer_weights = select_portfolio(ranked, metadata, args.top, gross, max_layer)
            turnover = sum(abs(new_weights.get(t, 0.0) - weights.get(t, 0.0)) for t in set(new_weights) | set(weights))
            turnover_total += turnover
            portfolio_value *= 1.0 - slippage * turnover
            weights = new_weights if not paused_until_rebalance else new_weights
            entry_prices = {
                ticker: price_on(closes_by_ticker, dates, ticker, di)
                for ticker in weights
            }
            paused_until_rebalance = False
            peak = max(peak, portfolio_value)
            holdings.append((day, list(weights), layer_weights))

    result = {
        "period_start": equity[0][0],
        "period_end": equity[-1][0],
        "days": len(equity),
        "strategy_x": equity[-1][1],
        "benchmark_x": benchmark_curve[-1][1],
        "strategy_cagr": cagr(equity),
        "benchmark_cagr": cagr(benchmark_curve),
        "strategy_maxdd": maxdd(equity),
        "benchmark_maxdd": maxdd(benchmark_curve),
        "strategy_sharpe": sharpe(equity),
        "benchmark_sharpe": sharpe(benchmark_curve),
        "turnover": turnover_total,
        "stops": stop_count,
        "circuits": circuit_count,
        "last_holdings": holdings[-1] if holdings else ("-", [], {}),
    }

    if quiet:
        return result

    print("=" * 92)
    print(f"BACKTEST  top-{args.top}, rebal={args.rebal}d, mom={args.mom_days}d")
    print(
        f"weights struct x{args.w_struct} mom x{args.w_mom} relstr x{args.w_rs} | "
        f"benchmark: {args.benchmark or 'equal-weight all'}"
    )
    print(
        f"cash reserve={args.cash_reserve:.0%}, max layer={args.max_layer_weight:.0%}, "
        f"slippage={args.slippage_bps:.1f} bps, stop={args.backtest_stop:.0%}, "
        f"portfolio stop={args.portfolio_stop:.0%}"
    )
    print("=" * 92)
    print(f"  period:        {result['period_start']} -> {result['period_end']} ({result['days']} days)")
    print(
        f"  strategy:      {result['strategy_x']:.2f}x  CAGR {result['strategy_cagr'] * 100:6.1f}%  "
        f"maxDD {result['strategy_maxdd'] * 100:6.1f}%  Sharpe {result['strategy_sharpe']:5.2f}"
    )
    print(
        f"  benchmark:     {result['benchmark_x']:.2f}x  CAGR {result['benchmark_cagr'] * 100:6.1f}%  "
        f"maxDD {result['benchmark_maxdd'] * 100:6.1f}%  Sharpe {result['benchmark_sharpe']:5.2f}"
    )
    edge = result["strategy_x"] - result["benchmark_x"]
    print(f"  vs benchmark:  {'+' if edge >= 0 else ''}{edge:.2f}x")
    print(f"  turnover:      {result['turnover']:.2f}x notional   stops: {result['stops']}   circuits: {result['circuits']}")
    last_day, last_names, layer_weights = result["last_holdings"]
    print(f"\n  Last rebalance ({last_day}) held: {', '.join(last_names) if last_names else '-'}")
    if layer_weights:
        print("  Layer weights:  " + ", ".join(f"{k[:18]}={v:.0%}" for k, v in sorted(layer_weights.items())))
    print("\n  In-sample unless paired with --sweep and held-out review. Not a profit promise.")
    print("  Idea-generation only. Not investment advice.")
    return result


def run_sweep(args):
    """Run a small robustness grid around the chosen defaults."""
    tops = sorted({max(2, args.top - 2), args.top, args.top + 2})
    rebals = sorted({21, args.rebal, 42, 63})
    moms = sorted({21, args.mom_days, 63, 126})
    weight_sets = [
        (1.0, 1.0, 1.0),
        (1.5, 1.0, 0.5),
        (1.0, 0.5, 1.0),
        (0.5, 1.0, 1.0),
    ]

    rows = []
    for top in tops:
        for rebal in rebals:
            for mom_days in moms:
                for ws, wm, wr in weight_sets:
                    trial = copy(args)
                    trial.top = top
                    trial.rebal = rebal
                    trial.mom_days = mom_days
                    trial.w_struct = ws
                    trial.w_mom = wm
                    trial.w_rs = wr
                    res = run_backtest(trial, quiet=True)
                    if res:
                        rows.append((res["strategy_x"] - res["benchmark_x"], trial, res))

    if not rows:
        print("No sweep results. Run _tools/fetch_history.py first.")
        return

    rows.sort(key=lambda x: x[0], reverse=True)
    wins = [x for x in rows if x[0] > 0]
    median_edge = sorted(x[0] for x in rows)[len(rows) // 2]
    print("=" * 100)
    print("ROBUSTNESS SWEEP  (prefer broad consistency over one best run)")
    print("=" * 100)
    print(f"runs={len(rows)}  win-rate-vs-benchmark={len(wins) / len(rows):.0%}  median edge={median_edge:+.2f}x")
    print(f"{'Edge':>8}{'StratX':>9}{'BenchX':>9}{'Top':>5}{'Rebal':>7}{'Mom':>6}{'Weights':>18}{'MaxDD':>9}  {'Last holdings'}")
    print("-" * 100)
    for edge, trial, res in rows[:12]:
        _, held, _ = res["last_holdings"]
        weights = f"{trial.w_struct:g}/{trial.w_mom:g}/{trial.w_rs:g}"
        print(
            f"{edge:>+8.2f}{res['strategy_x']:>9.2f}{res['benchmark_x']:>9.2f}"
            f"{trial.top:>5}{trial.rebal:>7}{trial.mom_days:>6}{weights:>18}"
            f"{res['strategy_maxdd'] * 100:>8.1f}%  {', '.join(held[:6])}"
        )
    print("\nIdea-generation only. Not investment advice.")


# ---------------- self-test ----------------
def selftest():
    passed = 0
    failed = 0

    def chk(name, condition):
        nonlocal passed, failed
        print(("  ok " if condition else "  XX ") + name)
        passed += int(bool(condition))
        failed += int(not condition)

    chk("zscore zeros on flat", all(v == 0 for v in zscore({"a": 5, "b": 5}).values()))
    z = zscore({"a": 1, "b": 2, "c": 3})
    chk("zscore symmetric", abs(z["a"] + z["c"]) < 1e-9 and abs(z["b"]) < 1e-9)
    chk("momentum simple", abs(momentum([100, 110], 1) - 0.10) < 1e-9)
    chk("momentum needs history", momentum([100], 5) is None)
    chk("sma", sma([1, 2, 3, 4], 2) == 3.5)
    lv = levels(100, 0.08, 0.16)
    chk("levels stop", lv["stop"] == 92.0)
    chk("levels target", lv["target"] == 116.0)
    chk("levels R", lv["r_multiple"] == 2.0)
    comp = composite({"a": 10, "b": 1, "c": 5}, {"a": 0.2, "b": -0.1, "c": 0.0}, {"a": 0.5, "b": -0.5, "c": 0.0}, 1, 1, 1)
    chk("composite ranks high-struct-high-mom first", max(comp, key=comp.get) == "a")
    weights, layers = select_portfolio(
        ["a", "b", "c"],
        {"a": {"layer_name": "L1"}, "b": {"layer_name": "L1"}, "c": {"layer_name": "L2"}},
        top=3,
        gross_exposure=0.9,
        max_layer_weight=0.35,
    )
    chk("layer cap skips crowded layer", list(weights) == ["a", "c"] and round(sum(weights.values()), 2) == 0.6)
    rs = percentile_scores({"a": 10, "b": 5, "c": 1})
    chk("relative strength ranks high positive", rs["a"] > rs["b"] > rs["c"])

    print(f"\nRESULT {passed} passed, {failed} failed")
    sys.exit(1 if failed else 0)


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--signals", action="store_true")
    parser.add_argument("--backtest", action="store_true")
    parser.add_argument("--sweep", action="store_true", help="run a robustness grid around the chosen parameters")
    parser.add_argument("--selftest", action="store_true")
    parser.add_argument("--top", type=int, default=5)
    parser.add_argument("--rebal", type=int, default=21)
    parser.add_argument("--mom-days", type=int, default=63, dest="mom_days")
    parser.add_argument("--w-struct", type=float, default=1.0, dest="w_struct")
    parser.add_argument("--w-mom", type=float, default=1.0, dest="w_mom")
    parser.add_argument("--w-rs", type=float, default=1.0, dest="w_rs")
    parser.add_argument("--stop", type=float, default=0.08, help="signal level stop fraction")
    parser.add_argument("--target", type=float, default=0.16, help="signal level target fraction")
    parser.add_argument("--backtest-stop", type=float, default=0.0, help="optional close-based stop-loss simulation")
    parser.add_argument("--portfolio-stop", type=float, default=0.0, help="optional portfolio drawdown circuit breaker")
    parser.add_argument("--cash-reserve", type=float, default=0.10)
    parser.add_argument("--max-layer-weight", type=float, default=0.35)
    parser.add_argument("--slippage-bps", type=float, default=10.0)
    parser.add_argument("--exclude", default="BOE,XFAB,SPCX",
                        help="comma-separated tickers to exclude from the strategy universe")
    parser.add_argument("--benchmark", default="")
    args = parser.parse_args()

    if args.selftest:
        return selftest()
    if not os.path.exists(HIST) and (args.signals or args.backtest or args.sweep):
        print("No history file. Run:  python3 _tools/fetch_history.py", file=sys.stderr)
        sys.exit(1)
    if args.signals:
        run_signals(args)
    elif args.backtest:
        run_backtest(args)
    elif args.sweep:
        run_sweep(args)
    else:
        print("Pick --signals, --backtest, --sweep, or --selftest. See --help.")


if __name__ == "__main__":
    main()
