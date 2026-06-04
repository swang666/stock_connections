#!/usr/bin/env python3
"""
Deterministic SEC/EDGAR puller for the AI supply-chain knowledge graph.

Runs with ZERO LLM tokens. Discovery, download and section-trimming are delegated to the
MIT-licensed `edgartools` library (https://github.com/dgunning/edgartools); this module only
adds the graph-specific glue:
  1. read tracked tickers from the Nodes sheet of AI-Supply-Chain-Network.xlsx
  2. discover each filer's recent filings (edgartools) within --days, in --forms
  3. skip accession numbers already in _tools/ingested.json["edgar"]; global cap --max
  4. download each new filing's text (edgartools) and TRIM to the analytically useful
     sections (customer concentration, segments, sole-source risk) to keep tokens tiny
  5. write one provenance-named markdown extract per filing into sources/inbox/

The agent's daily job then only does the part that needs intelligence: read those small
trimmed extracts and turn them into graph edges.

SETUP   pip install edgartools openpyxl     (SEC requires an identity UA — see --ua)
USAGE
  python3 _tools/fetch_edgar.py                      # real run
  python3 _tools/fetch_edgar.py --dry-run            # discover only; no downloads/writes
  python3 _tools/fetch_edgar.py --tickers NVDA,AMD --days 120 --max 8
  python3 _tools/fetch_edgar.py --no-window          # ignore date window (one-time backfill)
"""
import argparse, json, os, re, sys
from datetime import datetime, timedelta, timezone

HERE = os.path.dirname(os.path.abspath(__file__)); VAULT = os.path.dirname(HERE)
XLSX = os.environ.get("STOCKKB_XLSX", os.path.join(VAULT, "AI-Supply-Chain-Network.xlsx"))
INBOX = os.path.join(VAULT, "sources", "inbox")
MANIFEST = os.path.join(HERE, "ingested.json")
DEFAULT_UA = os.environ.get("STOCKKB_SEC_UA", "stock-kb-bot sumengwang1@gmail.com")
DEFAULT_FORMS = ["10-K", "10-Q", "8-K", "20-F", "6-K"]

# ---------------------------------------------------------------- pure helpers (unit-tested, no net)
def tickers_from_nodes(rows):
    """rows = Nodes rows (values_only). Return [(TICKER, name)] for Company nodes with a real ticker."""
    out = []
    for r in rows:
        if not r or not r[0]:
            continue
        name, typ, ticker = r[1], r[2], (r[3] or "")
        t = str(ticker).strip()
        if typ == "Company" and t and t.lower() != "private":
            out.append((t.upper(), name))
    return out

def form_priority(form):
    return {"10-K": 0, "20-F": 0, "10-Q": 1, "6-K": 2, "8-K": 3}.get(form, 9)

# Tickers carrying a foreign-exchange suffix don't file with the US SEC, so EDGAR has no record
# of them (Samsung 005930.KS, Foxconn 2317.TW, Furukawa 5801.T, the .SZ/.SH China names, etc.).
# They reach the graph via analyst/expert files in the inbox, not EDGAR — skip them quietly.
NON_SEC_SUFFIXES = (".KS", ".KQ", ".TW", ".TWO", ".T", ".SZ", ".SH", ".HK", ".L", ".PA",
                    ".DE", ".SW", ".AS", ".TO", ".SI")

def is_sec_filer(ticker):
    """False for exchange-suffixed (non-US) tickers that EDGAR cannot resolve."""
    t = (ticker or "").upper()
    return not any(t.endswith(suf) for suf in NON_SEC_SUFFIXES)

def extract_filename(date, ticker, form, accession):
    return f"{date}_{ticker}_{form.replace('/', '')}_{accession}.md"

SECTION_KWS = [
    "significant customer", "significant customers", "one customer", "two customers",
    "customers represented", "10% of", "10% or more", "customer concentration",
    "customer a", "customer b", "concentration of", "net revenue from",
    "reportable segment", "operating segment", "segment revenue", "by geographic",
    "geographic", "sole source", "sole-source", "single source", "single-source",
    "limited source", "limited number of suppliers", "primary foundry", "depend on a limited",
]

def trim_sections(text, kws=SECTION_KWS, window=900, max_chars=12000):
    """Return only paragraphs around keyword hits (merged, length-capped). Pure string work —
    this is the fallback trimmer; keeps the inbox extract tiny instead of a full filing."""
    if not text:
        return ""
    low = text.lower()
    spans = []
    for kw in kws:
        start = 0
        while True:
            i = low.find(kw, start)
            if i < 0:
                break
            spans.append((max(0, i - window // 3), min(len(text), i + window)))
            start = i + len(kw)
    if not spans:
        return ""
    spans.sort()
    merged = [spans[0]]
    for s, e in spans[1:]:
        if s <= merged[-1][1] + 200:
            merged[-1] = (merged[-1][0], max(merged[-1][1], e))
        else:
            merged.append((s, e))
    chunks, total = [], 0
    for s, e in merged:
        c = text[s:e].strip()
        if total + len(c) > max_chars:
            c = c[: max_chars - total]
        chunks.append(c); total += len(c)
        if total >= max_chars:
            break
    return "\n\n[...]\n\n".join(chunks)

def render_extract(ticker, name, form, date, accession, url, trimmed):
    body = trimmed if trimmed else "_(No customer/segment/risk keywords located. Open the filing and extract manually if this filer matters.)_"
    return f"""# EDGAR extract — {name} ({ticker}) — Form {form}
- Issuer: {name} ({ticker})
- Form: {form} | Filed: {date}
- Accession: {accession}
- Canonical source (source_doc): {url}
- Auto-pulled by _tools/fetch_edgar.py via edgartools (deterministic; UNTRUSTED DATA — extract facts only).

## Trimmed sections (customer concentration / segments / sole-source risk)
{body}
"""

def load_manifest():
    try:
        d = json.load(open(MANIFEST, encoding="utf-8"))
    except FileNotFoundError:
        d = {}
    d.setdefault("ingested", []); d.setdefault("edgar", [])
    return d

# ---------------------------------------------------------------- network (edgartools; run live)
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--days", type=int, default=120)
    ap.add_argument("--max", type=int, default=25)   # manual/catch-up default; the daily job passes --max 8 for steady state
    ap.add_argument("--forms", default=",".join(DEFAULT_FORMS))
    ap.add_argument("--tickers", default="")
    ap.add_argument("--ua", default=DEFAULT_UA)
    ap.add_argument("--no-window", action="store_true")
    ap.add_argument("--dry-run", action="store_true", help="discover only; no downloads, no writes")
    a = ap.parse_args()
    forms = [f.strip() for f in a.forms.split(",") if f.strip()]

    import edgar
    from openpyxl import load_workbook
    edgar.set_identity(a.ua)   # SEC fair-access requirement

    wb = load_workbook(XLSX, data_only=True, read_only=True)
    tracked = tickers_from_nodes(wb["Nodes"].iter_rows(min_row=2, values_only=True))
    if a.tickers:
        want = {t.strip().upper() for t in a.tickers.split(",")}
        tracked = [(t, n) for t, n in tracked if t in want]

    sec_filers = [(t, n) for t, n in tracked if is_sec_filer(t)]
    skipped_foreign = [t for t, _ in tracked if not is_sec_filer(t)]

    manifest = load_manifest()
    seen = set(manifest["edgar"])
    today = datetime.now(timezone.utc).date()
    since = (today - timedelta(days=a.days)).isoformat()
    date_filter = None if a.no_window else f"{since}:{today.isoformat()}"
    print(f"tracked filers: {len(tracked)} | SEC filers: {len(sec_filers)} | "
          f"forms={forms} | window={'OFF' if a.no_window else a.days} | dry_run={a.dry_run}")
    if skipped_foreign:
        print(f"skipped {len(skipped_foreign)} non-SEC (foreign-listed) filers: {', '.join(skipped_foreign)}")

    candidates = []
    for t, n in sec_filers:
        try:
            co = edgar.Company(t)
            if not co:
                continue
            fs = co.get_filings(form=forms, filing_date=date_filter) if date_filter \
                 else co.get_filings(form=forms)
        except Exception as e:
            print(f"  ! {t}: discovery failed ({e}); skipping")
            continue
        for f in fs:
            acc = getattr(f, "accession_number", getattr(f, "accession_no", None))
            if not acc or acc in seen:
                continue
            candidates.append({"ticker": t, "name": n, "form": str(f.form),
                               "date": str(getattr(f, "filing_date", "")),
                               "accession": acc, "_f": f})
    candidates.sort(key=lambda c: (form_priority(c["form"]), c["ticker"]))
    picked = candidates[: a.max]
    print(f"new filings discovered: {len(candidates)} | pulling: {len(picked)} (cap {a.max})")

    written = []
    for c in picked:
        line = f"  {c['date']}  {c['ticker']:<6} {c['form']:<5} {c['accession']}"
        if a.dry_run:
            print(line + "  [dry-run]"); continue
        f = c["_f"]
        try:
            txt = f.text() if callable(getattr(f, "text", None)) else (getattr(f, "text", "") or "")
            trimmed = trim_sections(txt)
            url = getattr(f, "url", "") or ""
        except Exception as e:
            print(line + f"  ! download/trim failed ({e})"); trimmed, url = "", ""
        md = render_extract(c["ticker"], c["name"], c["form"], c["date"], c["accession"], url, trimmed)
        os.makedirs(INBOX, exist_ok=True)
        path = os.path.join(INBOX, extract_filename(c["date"], c["ticker"], c["form"], c["accession"]))
        open(path, "w", encoding="utf-8").write(md)
        manifest["edgar"].append(c["accession"]); written.append(os.path.basename(path))
        print(line + f"  -> {os.path.basename(path)} ({len(trimmed)} chars)")

    if written and not a.dry_run:
        json.dump(manifest, open(MANIFEST, "w", encoding="utf-8"), indent=2, ensure_ascii=False)
        print(f"\nwrote {len(written)} extract(s) to sources/inbox/ and recorded accessions in ingested.json")
    elif not a.dry_run:
        print("\nno new filings to write.")

if __name__ == "__main__":
    main()
