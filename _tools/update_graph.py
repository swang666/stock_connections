#!/usr/bin/env python3
"""Regenerate everything derived from the live spreadsheet:
  1. AI Supply Chain Graph.html   -- interactive graph (catalyst + bottleneck modes)
  2. 00 - START HERE.md           -- the layer index (kept in sync with the node list)
  3. Degree (analytics) sheet     -- out/in/total degree per node

Bottleneck scores are computed in Python (bottleneck_analysis.py) and baked into node data so the
graph and the scoring engine never diverge. Run after any ingestion / manual edit."""
import os, json, sys, re
from collections import defaultdict
from openpyxl import load_workbook
HERE = os.path.dirname(os.path.abspath(__file__)); VAULT = os.path.dirname(HERE)
sys.path.insert(0, VAULT)
import bottleneck_analysis as ba
XLSX = os.environ.get("STOCKKB_XLSX", os.path.join(VAULT, "AI-Supply-Chain-Network.xlsx"))
TMPL = os.path.join(HERE, "graph_template_v3.html")
LAYERS = {1:"EDA & IP",2:"Semiconductor Equipment",3:"Foundry / Manufacturing",
 4:"Chip Designers (Accelerators/Networking)",5:"Memory (HBM/DRAM)",6:"Advanced Packaging / OSAT",
 7:"Optical & Interconnect",8:"Networking Systems",9:"Servers & ODM/OEM",10:"Hyperscalers / Cloud",
 11:"AI Labs / Model Developers",12:"Datacenter Infrastructure (REITs)",13:"Power & Cooling Equipment",
 14:"Power Generation / Utilities",0:"Demand Driver / Theme"}
LAYER_COLOR = {1:"7c3aed",2:"6d28d9",3:"2563eb",4:"059669",5:"0891b2",6:"0d9488",7:"0284c7",
 8:"4f46e5",9:"d97706",10:"dc2626",11:"db2777",12:"9333ea",13:"ea580c",14:"ca8a04",0:"475569"}

# ---------- read live data ----------
wb = load_workbook(XLSX, data_only=True, read_only=True)
nodes=[]
for r in wb["Nodes"].iter_rows(min_row=2, values_only=True):
    if not r or not r[0]: continue
    nodes.append({"id":r[0],"name":r[1],"type":r[2],"ticker":r[3] or "","layer":r[4],
                  "layer_name":r[5],"segment":r[6] or "","role":r[9] or ""})
def evtag(ev):
    """Derive a confidence tag from the evidence text: FACT (stated), VIEW (analyst opinion),
    ? (flagged/uncertain), or '' (untagged). Mirrors the FACT/VIEW discipline in CLAUDE.md."""
    t = (ev or "").lower()
    if "needs-" in t or t.strip().endswith("(?)") or " (?)" in t: return "?"
    has_f, has_v = "(fact" in t, "(view" in t
    if has_f and not has_v: return "FACT"
    if has_v and not has_f: return "VIEW"
    if has_f and has_v: return "FACT"   # mixed -> treat as fact-backed
    return ""
def public_src(s):
    """Only expose a source link if it's a public http(s) URL. Private/local source_doc values
    (analyst PDFs, expert-call docs, X-scan files) are NOT in the repo, so linking them would be a
    broken relative link AND would leak a private filename onto the public page — drop those."""
    s = str(s or "").strip()
    return s if s.lower().startswith(("http://", "https://")) else ""
edges=[]
for r in wb["Edges"].iter_rows(min_row=2, values_only=True):
    if not r or not r[0]: continue
    ev = r[6] or ""
    edges.append({"s":r[0],"r":r[2],"t":r[3],"w":r[5] if isinstance(r[5],(int,float)) else 2,
                  "ev":ev, "src":public_src(r[8]), "asof":(str(r[7])[:10] if r[7] else ""),
                  "tag":evtag(ev)})
# node extras: country + key_products (already-present columns, just not previously surfaced)
node_extra={}
for r in wb["Nodes"].iter_rows(min_row=2, values_only=True):
    if r and r[0]: node_extra[r[0]] = {"country": r[7] or "", "key_products": r[8] or ""}
wb.close()
for nd in nodes:
    nd.update(node_extra.get(nd["id"], {"country":"", "key_products":""}))

# ---------- 1) interactive graph ----------
bn, be = ba.load(); rows = ba.analyze(bn, be); S = {x["id"]: x for x in rows}
for nd in nodes:
    x = S.get(nd["id"])
    if x:
        nd.update({"bscore":x["score"],"breach":x["reach"],"balts":x["alts"],
                   "bconstr":x["constraint"],"bwhy":x["why"],"bemerging":bool(x["emerging"])})
    else:
        nd.update({"bscore":0,"breach":0,"balts":0,"bconstr":0,"bwhy":"","bemerging":False})
lcolor={str(k):"#"+v for k,v in LAYER_COLOR.items()}; lname={str(k):v for k,v in LAYERS.items()}
# price proxy config (optional; empty PROXY_URL just disables the live-price layer)
try:
    pricecfg=json.load(open(os.path.join(HERE,"price_config.json"),encoding="utf-8"))
    pricecfg={"PROXY_URL":pricecfg.get("PROXY_URL","") or "","REFRESH_SECONDS":pricecfg.get("REFRESH_SECONDS",60)}
except Exception:
    pricecfg={"PROXY_URL":"","REFRESH_SECONDS":60}
# catalyst calendar (optional; missing file -> empty calendar)
try:
    catalysts=json.load(open(os.path.join(HERE,"catalysts.json"),encoding="utf-8"))
except Exception:
    catalysts=[]
tmpl=open(TMPL,encoding="utf-8").read()
out=(tmpl.replace("/*NODES*/","NODES = "+json.dumps(nodes))
        .replace("/*EDGES*/","EDGES = "+json.dumps(edges))
        .replace("/*LCOLOR*/","LCOLOR = "+json.dumps(lcolor))
        .replace("/*LNAME*/","LNAME = "+json.dumps(lname))
        .replace("/*PRICECFG*/","PRICECFG = "+json.dumps(pricecfg))
        .replace("/*CATALYSTS*/","CATALYSTS = "+json.dumps(catalysts)))
open(os.path.join(VAULT,"AI Supply Chain Graph.html"),"w",encoding="utf-8").write(out)
# also publish a copy as index.html at the repo root so GitHub Pages serves it at the site root
# (share link becomes https://<user>.github.io/<repo>/ with no %20-escaped filename).
open(os.path.join(VAULT,"index.html"),"w",encoding="utf-8").write(out)

# ---------- 2) regenerate START HERE layer index ----------
by_layer=defaultdict(list)
for nd in nodes:
    by_layer[nd["layer"]].append(nd["name"])
def links(names): return ", ".join(f"[[{n}]]" for n in sorted(names))
idx=["## Layers (upstream → downstream)",""]
for L in range(1,15):
    if by_layer.get(L):
        idx.append(f"**L{L} — {LAYERS[L]}**"); idx.append(links(by_layer[L])); idx.append("")
idx.append("## Demand drivers (themes)")
idx.append(links(by_layer.get(0,[])) or "_none_"); idx.append("")
SH=os.path.join(VAULT,"00 - START HERE.md")
txt=open(SH,encoding="utf-8").read()
marker="## Layers (upstream → downstream)"
head=txt.split(marker)[0].rstrip()+"\n\n"
open(SH,"w",encoding="utf-8").write(head+"\n".join(idx).rstrip()+"\n")

# ---------- 3) populate Degree (analytics) sheet ----------
outd=defaultdict(int); ind=defaultdict(int)
for e in edges:
    outd[e["s"]]+=1; ind[e["t"]]+=1
id2name={nd["id"]:nd["name"] for nd in nodes}; id2layer={nd["id"]:nd["layer_name"] for nd in nodes}
wb2=load_workbook(XLSX)            # write mode
if "Degree (analytics)" in wb2.sheetnames:
    ws=wb2["Degree (analytics)"]; wb2.remove(ws)
ws=wb2.create_sheet("Degree (analytics)")
ws.append(["node_id","name","layer_name","out_degree","in_degree","total_connections"])
for nd in sorted(nodes, key=lambda n:-(outd[n["id"]]+ind[n["id"]])):
    i=nd["id"]; ws.append([i,id2name[i],id2layer[i],outd[i],ind[i],outd[i]+ind[i]])
wb2.save(XLSX)

print(f"regenerated: graph ({len(nodes)} nodes, {len(edges)} edges) | START HERE | Degree sheet | prices+catalysts baked")
