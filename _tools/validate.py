#!/usr/bin/env python3
"""
Integrity checks for the AI supply-chain knowledge graph. Run after any ingestion / manual edit
(and as the last step of the daily job) to catch regressions before they reach the graph.

  python3 _tools/validate.py            # human-readable report, exit 1 if any ERROR
  python3 _tools/validate.py --strict   # also treat WARNINGs as failures

Checks:
  ERROR  dangling edges            -- edge endpoint missing from Nodes
  ERROR  duplicate node_id
  ERROR  bad strength (not 1..3)
  ERROR  broken [[wiki-links]]     -- link target has no matching page
  ERROR  node<->page mismatch      -- Company/Theme node without a page, or page without a node
  WARN   orphan nodes              -- node with no edges
  WARN   evidence missing          -- edge with no evidence text
  WARN   evidence untagged         -- evidence present but no (FACT/VIEW ...) tag
  WARN   source_doc missing
  WARN   flagged edges             -- evidence contains NEEDS- (awaiting a primary source)
"""
import os, re, sys, glob, argparse
from collections import defaultdict, Counter
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__)); VAULT = os.path.dirname(HERE)
XLSX = os.environ.get("STOCKKB_XLSX", os.path.join(VAULT, "AI-Supply-Chain-Network.xlsx"))

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--strict", action="store_true")
    a = ap.parse_args()
    errors, warns = [], []

    wb = load_workbook(XLSX, data_only=True, read_only=True)
    nodes, dup = {}, []
    ntype = {}
    for r in wb["Nodes"].iter_rows(min_row=2, values_only=True):
        if not r or not r[0]: continue
        if r[0] in nodes: dup.append(r[0])
        nodes[r[0]] = r[1]; ntype[r[0]] = r[2]
    edges = [r for r in wb["Edges"].iter_rows(min_row=2, values_only=True) if r and r[0]]

    for d in dup: errors.append(f"duplicate node_id: {d}")

    deg = defaultdict(int)
    for e in edges:
        s, t, w, ev, src = e[0], e[3], e[5], e[6], e[8]
        if s not in nodes: errors.append(f"dangling edge source: {s} ({e[1]})")
        if t not in nodes: errors.append(f"dangling edge target: {t} ({e[4]})")
        if s in nodes and t in nodes: deg[s]+=1; deg[t]+=1
        if not (isinstance(w,(int,float)) and 1 <= w <= 3):
            errors.append(f"bad strength {w!r}: {e[1]} -> {e[4]}")
        label = f"{e[1]} --{e[2]}--> {e[4]}"
        if not ev: warns.append(f"evidence missing: {label}")
        else:
            evl = str(ev).lower()
            if "needs-" in evl: warns.append(f"flagged (awaiting source): {label}")
            elif "(fact" not in evl and "(view" not in evl:
                warns.append(f"evidence untagged (no FACT/VIEW): {label}")
        if not src: warns.append(f"source_doc missing: {label}")

    for nid, nm in nodes.items():
        if deg[nid] == 0: warns.append(f"orphan node (no edges): {nm}")

    # node <-> page parity + broken wiki-links
    comp_pages = {os.path.splitext(os.path.basename(p))[0] for p in glob.glob(os.path.join(VAULT,"Companies","*.md"))}
    theme_pages = {os.path.splitext(os.path.basename(p))[0] for p in glob.glob(os.path.join(VAULT,"Themes","*.md"))}
    all_pages = comp_pages | theme_pages
    for nid, nm in nodes.items():
        if ntype.get(nid) == "Company" and nm not in comp_pages:
            errors.append(f"Company node without page: {nm}")
        if ntype.get(nid) == "Theme" and nm not in theme_pages:
            errors.append(f"Theme node without page: {nm}")
    node_names = set(nodes.values())
    for p in comp_pages:
        if p not in node_names: errors.append(f"Company page without node: {p}")
    for p in theme_pages:
        if p not in node_names: errors.append(f"Theme page without node: {p}")
    for path in glob.glob(os.path.join(VAULT,"Companies","*.md"))+glob.glob(os.path.join(VAULT,"Themes","*.md")):
        txt = open(path, encoding="utf-8").read()
        for m in re.findall(r"\[\[([^\]|#]+)(?:[|#][^\]]*)?\]\]", txt):
            if m.strip() not in all_pages:
                errors.append(f"broken wiki-link in {os.path.basename(path)}: [[{m.strip()}]]")

    wc = Counter(w.split(":")[0].split(" (")[0] for w in warns)
    print("="*70); print("KNOWLEDGE-GRAPH VALIDATION"); print("="*70)
    print(f"nodes={len(nodes)}  edges={len(edges)}  errors={len(errors)}  warnings={len(warns)}")
    if warns:
        print("\nWARNINGS by type:")
        for k,v in wc.most_common(): print(f"  {v:>3}  {k}")
    if errors:
        print("\nERRORS:")
        for e in errors[:50]: print("  x", e)
    print("\n" + ("FAIL" if errors or (a.strict and warns) else "PASS"))
    sys.exit(1 if errors or (a.strict and warns) else 0)

if __name__ == "__main__":
    main()
