#!/usr/bin/env python3
"""
Daily historical price bars for the quant model.

The primary output is:

    _tools/prices_history.csv     columns: date,ticker,open,high,low,close,volume

Stooq is tried first because it is simple CSV. If Stooq returns a browser challenge or empty data,
the script falls back to Yahoo's public v8/chart endpoint. Plain US symbols are supported; foreign
exchange-suffixed tickers and PRIVATE placeholders are skipped.
"""
import argparse
import csv
import io
import json
import os
import sys
import time
import urllib.request
from datetime import date, datetime, timezone

HERE = os.path.dirname(os.path.abspath(__file__))
VAULT = os.path.dirname(HERE)
XLSX = os.environ.get("STOCKKB_XLSX", os.path.join(VAULT, "AI-Supply-Chain-Network.xlsx"))
OUT = os.path.join(HERE, "prices_history.csv")
NON_US_SUFFIXES = (
    ".KS", ".KQ", ".TW", ".TWO", ".T", ".SZ", ".SH", ".SS", ".HK", ".L", ".PA",
    ".DE", ".SW", ".AS", ".TO", ".SI", ".ST",
)
DEFAULT_EXTRAS = ["NVDA", "SOXX", "SMH", "QQQ", "SPY"]


def graph_tickers():
    from openpyxl import load_workbook

    wb = load_workbook(XLSX, data_only=True, read_only=True)
    out = []
    for row in wb["Nodes"].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        ticker = str(row[3] or "").strip().upper()
        if row[2] == "Company" and is_us(ticker):
            out.append(ticker)
    return out


def is_us(ticker):
    ticker = ticker.upper().strip()
    if not ticker or "PRIVATE" in ticker or ticker.startswith("(") or ticker.endswith(")"):
        return False
    if any(ticker.endswith(suffix) for suffix in NON_US_SUFFIXES):
        return False
    return all(ch.isalnum() or ch in ".-" for ch in ticker)


def stooq_symbol(ticker):
    return ticker.lower() + ".us"


def fetch_one_stooq(ticker, start, ua):
    url = f"https://stooq.com/q/d/l/?s={stooq_symbol(ticker)}&d1={start}&d2={date.today():%Y%m%d}&i=d"
    req = urllib.request.Request(url, headers={"User-Agent": ua})
    with urllib.request.urlopen(req, timeout=30) as resp:
        txt = resp.read().decode("utf-8", "replace")
    if "<html" in txt.lower() or "javascript" in txt.lower():
        return []

    rows = list(csv.DictReader(io.StringIO(txt)))
    out = []
    for row in rows:
        if "Date" not in row or not row.get("Close"):
            continue
        try:
            out.append({
                "date": row["Date"],
                "ticker": ticker,
                "open": row["Open"],
                "high": row["High"],
                "low": row["Low"],
                "close": row["Close"],
                "volume": row.get("Volume", ""),
            })
        except Exception:
            continue
    return out


def yahoo_epoch(yyyymmdd):
    dt = datetime.strptime(yyyymmdd, "%Y%m%d").replace(tzinfo=timezone.utc)
    return int(dt.timestamp())


def fetch_one_yahoo(ticker, start, ua):
    p1 = yahoo_epoch(start)
    p2 = int(time.time()) + 86400
    url = (
        f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}"
        f"?period1={p1}&period2={p2}&interval=1d&events=history&includeAdjustedClose=true"
    )
    req = urllib.request.Request(url, headers={"User-Agent": ua})
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = json.loads(resp.read().decode("utf-8", "replace"))

    result = (data.get("chart", {}).get("result") or [None])[0]
    if not result:
        return []
    timestamps = result.get("timestamp") or []
    quote = (result.get("indicators", {}).get("quote") or [{}])[0]
    adj = (result.get("indicators", {}).get("adjclose") or [{}])[0].get("adjclose") or []

    out = []
    for i, ts in enumerate(timestamps):
        try:
            raw_close = quote.get("close", [None])[i]
            close = adj[i] if i < len(adj) and adj[i] is not None else raw_close
            if close is None:
                continue
            out.append({
                "date": datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d"),
                "ticker": ticker,
                "open": quote.get("open", [None])[i],
                "high": quote.get("high", [None])[i],
                "low": quote.get("low", [None])[i],
                "close": close,
                "volume": quote.get("volume", [None])[i],
            })
        except (IndexError, KeyError, TypeError):
            continue
    return out


def fetch_one(ticker, start, ua):
    rows = fetch_one_stooq(ticker, start, ua)
    if rows:
        return rows, "stooq"
    rows = fetch_one_yahoo(ticker, start, ua)
    return rows, "yahoo" if rows else "none"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--tickers", default="")
    parser.add_argument("--extra-tickers", default=",".join(DEFAULT_EXTRAS),
                        help="benchmark/watch tickers to append when --tickers is not supplied")
    parser.add_argument("--years", type=int, default=3)
    parser.add_argument("--out", default=OUT)
    parser.add_argument("--ua", default="Mozilla/5.0 stock-kb-quant")
    parser.add_argument("--throttle", type=float, default=0.3)
    args = parser.parse_args()

    explicit = [t.strip().upper() for t in args.tickers.split(",") if t.strip()]
    tickers = explicit or graph_tickers()
    if not explicit:
        tickers += [t.strip().upper() for t in args.extra_tickers.split(",") if t.strip()]
    tickers = list(dict.fromkeys(t for t in tickers if is_us(t)))

    start = f"{date.today().year - args.years}{date.today():%m%d}"
    print(f"fetching {len(tickers)} US tickers, ~{args.years}y...")
    allrows = []
    ok = 0
    for ticker in tickers:
        try:
            rows, source = fetch_one(ticker, start, args.ua)
            time.sleep(args.throttle)
            if rows:
                allrows += rows
                ok += 1
                print(f"  {ticker}: {len(rows)} bars ({source})")
            else:
                print(f"  {ticker}: no data, skipped")
        except Exception as exc:
            print(f"  {ticker}: failed ({exc})")

    if not allrows:
        print("no data fetched.")
        sys.exit(1)
    with open(args.out, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["date", "ticker", "open", "high", "low", "close", "volume"])
        writer.writeheader()
        writer.writerows(allrows)
    print(f"wrote {len(allrows)} rows for {ok} tickers -> {args.out}")


if __name__ == "__main__":
    main()
