#!/usr/bin/env python3
"""
Bottleneck finder for the AI supply-chain knowledge graph.

A "bottleneck" = a node the rest of the chain depends on, with few substitutes, that is
supply-constrained or has pricing power. Those are where margin and pricing accrue — i.e. where
the investment opportunities concentrate. This scores every company node on three signals,
all derived from the graph + the evidence text you already have:

  1. REACH (criticality)  — how much of the chain sits downstream of it (transitive dependents).
  2. UNIQUENESS (the choke) — how few alternative suppliers do the same job into the same buyers.
  3. CONSTRAINT (tightness) — shortage / sole-source / pricing-power language in the edge evidence
                              and the company's role text (sole, monopoly, bottleneck, tight, ASP,
                              LTA, EUV, CoWoS, HBM, allocation, lead time, undersupplied ...).

  bottleneck score = 100 · sqrt(reach) · (0.35 + 0.65·uniqueness) · (0.55 + 0.45·constraint)
                          · (0.9 + 0.1·demand_pull) · confidence
  confidence ∈ [0.7,1.0] discounts nodes whose evidence is mostly analyst VIEW vs stated FACT.

Reads the LIVE spreadsheet, so anything you ingest later automatically feeds the ranking.
This is idea-generation, not investment advice — verify every chokepoint against primary sources.

USAGE
  python3 bottleneck_analysis.py                 # ranked table, all layers
  python3 bottleneck_analysis.py --top 25 --out bottlenecks.csv
  python3 bottleneck_analysis.py --layer 1-7     # restrict to upstream layers
"""
import argparse, csv, os, re
from collections import defaultdict, deque
from openpyxl import load_workbook

# Resolve the spreadsheet relative to this script so the engine works regardless of where it
# runs (sandbox, Windows/Obsidian, CI). Override with the STOCKKB_XLSX env var if needed.
HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.environ.get("STOCKKB_XLSX", os.path.join(HERE, "AI-Supply-Chain-Network.xlsx"))

def bucket(rel):
    r = rel.lower()
    # NOTE: order matters and checks are intentionally specific. "manufactures chips" must be
    # tested before the EDA/IP rule, because the loose substring "ip" matches inside "chips".
    if "manufactures chips" in r: return "foundry"
    if "eda" in r or "cpu ip" in r or " ip " in r or r.endswith(" ip") or "licenses" in r: return "eda_ip"
    if "euv" in r or "equipment" in r or "etch" in r or "deposition" in r or "inspection" in r: return "equipment"
    if "cowos" in r or "packaging" in r or "osat" in r: return "packaging"
    if "glass" in r or "substrate" in r: return "glass_substrate"  # glass-core substrate makers are mutual alternatives
    if "hbm" in r or "memory" in r or "dram" in r: return "memory"
    if "gpu" in r or "custom silicon" in r or "asic" in r: return "accelerator"
    # Optical FIBER (the physical cable plant) is NOT a substitute for optical modules/DSP/
    # transceivers — different layers of the optical stack. Bucket fiber separately so a fiber
    # supplier (Corning) isn't counted as competing with module/DSP makers (Marvell, AOE) into a
    # shared customer. (Fix 2026-06-17: this conflation crushed Corning's uniqueness 1.0→0.33 and
    # halved its bottleneck score on the day the AWS fiber deal — a bullish FACT — was added.)
    if "fiber" in r: return "optical_fiber"
    if "optical" in r or "module" in r: return "optical"
    if "switch" in r or "networking" in r: return "networking"
    if "cooling" in r: return "power_cooling"
    if "power" in r or "turbine" in r or "grid" in r or "nuclear" in r: return "power_gen"
    if "electrical" in r: return "electrical"
    if "server" in r or "assemble" in r: return "servers"
    if "datacenter" in r: return "datacenter"
    if "runs compute" in r: return "cloud"
    return r

def channel(rel):
    r = rel.lower()
    if "compete" in r: return "compete"
    if "benefits from" in r: return "benefit"
    if "drives" in r and "demand" in r: return "driver"
    if "runs compute on" in r: return "compute"
    return "supply"

KW_STRONG = ["bottleneck","sole","monopoly","constrained","shortage","undersupplied","scarce",
             "gatekeeper","near-monopoly","sole-source","only foundry","only supplier","tight supply"]
KW_MED = ["capacity","lead time","lead-time","allocation","lta","locked","premium","asp","pricing",
          "price increase","undersupply","tight","euv","cowos","hbm","vast majority","primary foundry"]

def load():
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    nodes = {}
    for r in wb["Nodes"].iter_rows(min_row=2, values_only=True):
        if not r or not r[0]: continue
        nodes[r[0]] = {"name":r[1],"type":r[2],"layer":r[4],"layer_name":r[5],
                       "segment":r[6] or "","role":(r[9] or "")}
    edges = []
    for r in wb["Edges"].iter_rows(min_row=2, values_only=True):
        if not r or not r[0]: continue
        edges.append({"s":r[0],"rel":r[2] or "","t":r[3],
                      "w":r[5] if isinstance(r[5],(int,float)) else 2,"ev":r[6] or ""})
    return nodes, edges

def analyze(nodes, edges):
    # supply graph up->down (normalize compute so provider is upstream)
    down = defaultdict(set)               # up -> {downstream}
    supply_edges = []                     # (up, down, bucket)
    theme_inc = defaultdict(int)          # node -> # theme edges incident
    compete = defaultdict(set)
    ev_by_src = defaultdict(list)         # supplier node -> [evidence text]
    for e in edges:
        ch = channel(e["rel"])
        if ch in ("supply","compute"):
            up, dn = (e["s"], e["t"]) if ch == "supply" else (e["t"], e["s"])
            down[up].add(dn); supply_edges.append((up, dn, bucket(e["rel"])))
            ev_by_src[up].append(e["ev"])
        elif ch in ("benefit","driver"):
            theme_inc[e["s"]] += 1; theme_inc[e["t"]] += 1
            ev_by_src[e["s"]].append(e["ev"])
        elif ch == "compete":
            compete[e["s"]].add(e["t"]); compete[e["t"]].add(e["s"])

    # 1) reach = distinct transitive descendants
    def reach(n):
        seen=set(); q=deque(down[n])
        while q:
            x=q.popleft()
            if x in seen: continue
            seen.add(x); q.extend(down[x]-seen)
        return len(seen)
    reach_raw = {n: reach(n) for n in nodes}
    maxreach = max(reach_raw.values()) or 1

    # providers of (customer, bucket)
    prov = defaultdict(set)
    for up, dn, bk in supply_edges:
        prov[(dn, bk)].add(up)
    competitors = defaultdict(set)
    for up, dn, bk in supply_edges:
        competitors[up] |= (prov[(dn, bk)] - {up})
    for n in nodes:
        competitors[n] |= compete[n]

    # 3) constraint from evidence + role text
    def constraint(n):
        txt = " ".join(ev_by_src.get(n, [])) + " " + nodes[n]["role"]
        tl = txt.lower(); wt = 0
        for k in KW_STRONG:
            if k in tl: wt += 2
        for k in KW_MED:
            if k in tl: wt += 1
        emerging = "(view" in tl and "(fact" not in tl
        # capture a short reason phrase
        hit = next((k for k in KW_STRONG+KW_MED if k in tl), "")
        return min(1.0, wt/4.0), emerging, hit

    # confidence: discount nodes whose evidence is mostly analyst VIEW vs stated FACT.
    def confidence(n):
        txt = " ".join(ev_by_src.get(n, [])).lower()
        nf, nv = txt.count("(fact"), txt.count("(view")
        if nf + nv == 0:
            return 1.0                     # untagged -> no opinion, don't penalise
        return 0.7 + 0.3 * (nf / (nf + nv))

    maxtheme = max(theme_inc.values()) if theme_inc else 1
    rows=[]
    for n, meta in nodes.items():
        if meta["type"] == "Theme" or not meta["layer"]:  # skip themes
            continue
        rn = reach_raw[n]/maxreach
        uniq = 1.0/(1.0+len(competitors[n]))
        constr, emerging, hit = constraint(n)
        demand = theme_inc.get(n,0)/maxtheme
        conf = confidence(n)
        score = 100*(rn**0.5)*(0.35+0.65*uniq)*(0.55+0.45*constr)*(0.9+0.1*demand)*conf
        rows.append({"id":n,"name":meta["name"],"layer":meta["layer"],
                     "layer_name":meta["layer_name"],"score":round(score,1),
                     "reach":reach_raw[n],"alts":len(competitors[n]),
                     "uniq":round(uniq,2),"constraint":round(constr,2),"conf":round(conf,2),
                     "emerging":emerging,"why":hit})
    rows.sort(key=lambda x:-x["score"])
    return rows

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--top", type=int, default=20)
    ap.add_argument("--out"); ap.add_argument("--layer")
    a = ap.parse_args()
    nodes, edges = load()
    rows = analyze(nodes, edges)
    if a.layer:
        lo,hi = (a.layer.split("-")+[a.layer])[:2]; lo,hi=int(lo),int(hi)
        rows = [r for r in rows if lo<=r["layer"]<=hi]
    print("="*94)
    print("AI SUPPLY-CHAIN BOTTLENECKS  (higher score = more critical + harder to substitute + tighter)")
    print("="*94)
    print(f"{'#':<3}{'Score':>6}  {'Company':<26}{'Layer':<26}{'Reach':>6}{'Alts':>5}  Constraint / why")
    print("-"*94)
    for i,r in enumerate(rows[:a.top],1):
        lvl = "HIGH" if r["constraint"]>=0.75 else ("MED" if r["constraint"]>=0.4 else "low")
        tag = " *emerging" if r["emerging"] else ""
        print(f"{i:<3}{r['score']:>6.1f}  {r['name'][:24]:<26}{r['layer_name'][:24]:<26}"
              f"{r['reach']:>6}{r['alts']:>5}  {lvl} ({r['why']}){tag}")
    if a.out:
        with open(a.out,"w",newline="",encoding="utf-8") as f:
            w=csv.DictWriter(f, fieldnames=["score","name","layer_name","reach","alts","uniq","constraint","conf","emerging","why"])
            w.writeheader()
            for r in rows: w.writerow({k:r[k] for k in w.fieldnames})
        print("\n-> wrote", a.out)
    print("\nReach = # downstream dependents · Alts = # substitute suppliers (0 = sole-source)")
    print("Idea-generation only; verify against filings. Not investment advice.")

if __name__ == "__main__":
    main()
