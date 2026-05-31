#!/usr/bin/env python3
"""
Catalyst / chain-effect analysis over the AI supply-chain knowledge graph.

Give it a seed event (which company/theme is hit, whether it's good or bad, and whether it's a
demand-side or supply-side shock). It propagates the shock through the relationship graph and
ranks every other company by how much it is likely to move -- surfacing the non-obvious 2nd/3rd
order beneficiaries and losers.

Reads the live data from  AI-Supply-Chain-Network.xlsx  (Nodes + Edges sheets), so any edges you
add later (e.g. from analyst reports) automatically feed the analysis.

--------------------------------------------------------------------------------------------------
USAGE
  python3 catalyst_analysis.py --seeds NVDA --polarity pos --kind demand
  python3 catalyst_analysis.py --seeds "TSMC" --polarity neg --kind supply --hops 3
  python3 catalyst_analysis.py --seeds T_POWER --polarity pos --kind macro
  python3 catalyst_analysis.py            # runs three built-in demo catalysts

  --seeds     comma-separated node ids OR names (e.g. NVDA  or  "NVIDIA,AMD")
  --polarity  pos | neg          (is the event good or bad for the seed?)
  --kind      demand | supply | macro
  --hops      max chain length (default 3)
  --decay     per-hop fade 0-1   (default 0.5)
  --out       write ranked results to this CSV

MODEL (heuristic, fully transparent -- see the proposal doc for the rationale)
  * DEMAND-side shock flows UPSTREAM  (seed -> its suppliers); SUPPLY-side flows DOWNSTREAM.
  * Sign carries along the chain unchanged, except 'competes with' flips it.
  * A company that 'drives demand for' a theme spills over to other companies that 'benefit from'
    that same theme (this is how a GPU demand surge reaches memory & packaging names).
  * Impact fades by `decay` each hop and scales by edge strength (1-3).
This is idea-generation, not investment advice. Verify every chain against primary sources.
--------------------------------------------------------------------------------------------------
"""
import argparse, csv, os, sys
from collections import defaultdict, deque
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.environ.get("STOCKKB_XLSX", os.path.join(HERE, "AI-Supply-Chain-Network.xlsx"))

# ---------- load graph ----------
def load_graph():
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    nodes = {}
    for row in wb["Nodes"].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        nid, name, typ, tic, lay, layname = row[0], row[1], row[2], row[3], row[4], row[5]
        nodes[nid] = {"name": name, "type": typ, "ticker": tic, "layer": lay, "layer_name": layname}
    edges = []
    for row in wb["Edges"].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        edges.append({"s": row[0], "rel": (row[2] or "").lower(), "t": row[3],
                      "w": row[5] if isinstance(row[5], (int, float)) else 2,
                      "ev": (row[6] or "")})
    return nodes, edges

def edge_conf(ev):
    """Down-weight edges supported only by analyst VIEW vs stated FACT. Untagged -> neutral."""
    t = (ev or "").lower()
    f, v = "(fact" in t, "(view" in t
    if v and not f:
        return 0.6
    return 1.0

def classify(rel):
    if "compete" in rel:                 return "compete"
    if "benefits from" in rel:           return "benefit"   # company -> theme
    if "drives" in rel and "demand" in rel: return "driver" # company -> theme
    if "runs compute on" in rel:         return "compute"   # customer -> provider
    return "supply"                                          # supplier -> customer

# ---------- build influence links ----------
# links[a] = list of (b, sign, factor, channel)  meaning "a up  =>  b moves <sign> by factor"
def build_links(nodes, edges):
    links = defaultdict(list)
    for e in edges:
        s, t, w = e["s"], e["t"], e["w"]
        if s not in nodes or t not in nodes:
            continue
        f = (w / 3.0) * edge_conf(e.get("ev"))
        ch = classify(e["rel"])
        if ch in ("supply", "compute"):
            up, down = (s, t) if ch == "supply" else (t, s)  # compute: provider(t) is upstream
            links[down].append((up,  1, f, "demand"))   # customer demand lifts supplier
            links[up].append((down, 1, f, "supply"))    # supplier health lifts customer
        elif ch == "benefit":            # company s benefits from theme t  => theme drives company
            links[t].append((s, 1, f, "thematic"))
        elif ch == "driver":             # company s drives demand for theme t
            links[s].append((t, 1, f, "driver"))
        elif ch == "compete":
            links[s].append((t, -1, f * 0.6, "compete"))
            links[t].append((s, -1, f * 0.6, "compete"))
    return links

ACTIVE = {
    "demand": {"demand", "thematic", "driver", "compete"},
    "supply": {"supply", "thematic", "compete"},
    "macro":  {"thematic", "driver", "demand", "compete"},
}
CH_LABEL = {"demand": "demand pull (upstream)", "supply": "supply push (downstream)",
            "thematic": "theme spillover", "driver": "drives theme", "compete": "competitive (inverse)"}

# ---------- propagate ----------
def propagate(seeds, polarity, kind, links, hops=3, decay=0.5, eps=0.04):
    active = ACTIVE[kind]
    impact = defaultdict(float)
    best = {}   # node -> (abs_contrib, path, channel)
    seedset = set(seeds)
    q = deque((s, float(polarity), 0, [s], None) for s in seeds)
    guard = 0
    while q and guard < 200000:
        guard += 1
        node, val, hop, path, _ = q.popleft()
        if hop >= hops:
            continue
        for (b, sign, factor, ch) in links.get(node, []):
            if ch not in active:
                continue
            contrib = val * sign * factor * decay
            if abs(contrib) < eps or b in seedset:
                continue
            impact[b] += contrib
            if b not in best or abs(contrib) > best[b][0]:
                best[b] = (abs(contrib), path + [b], ch)
            q.append((b, contrib, hop + 1, path + [b], ch))
    return impact, best

# ---------- report ----------
def resolve(seeds_arg, nodes):
    name2id = {v["name"].lower(): k for k, v in nodes.items()}
    out = []
    for tok in seeds_arg.split(","):
        tok = tok.strip()
        if not tok:
            continue
        if tok in nodes:
            out.append(tok)
        elif tok.lower() in name2id:
            out.append(name2id[tok.lower()])
        else:
            # ticker match
            m = [k for k, v in nodes.items() if (v["ticker"] or "").lower() == tok.lower()]
            if m:
                out.append(m[0])
            else:
                print(f"  ! seed '{tok}' not found -- skipping", file=sys.stderr)
    return out

def run(seeds, polarity, kind, nodes, links, hops, decay, out_csv=None, label=""):
    impact, best = propagate(seeds, polarity, kind, links, hops, decay)
    ranked = sorted(impact.items(), key=lambda kv: -abs(kv[1]))
    seed_names = ", ".join(nodes[s]["name"] for s in seeds)
    print("\n" + "=" * 86)
    print(f"CATALYST{(' — ' + label) if label else ''}")
    print(f"  Seed: {seed_names}   |   Polarity: {'POSITIVE' if polarity>0 else 'NEGATIVE'}   |   Type: {kind.upper()}")
    print("=" * 86)
    print(f"{'Rank':<5}{'Impact':>8}  {'Dir':<4}{'Hops':>5}  {'Company':<26}{'Path / mechanism'}")
    print("-" * 86)
    rows = []
    for i, (nid, val) in enumerate(ranked[:18], 1):
        if abs(val) < 0.04:
            continue
        path = best.get(nid, (0, [nid], ""))[1]
        ch = best.get(nid, (0, [], ""))[2]
        hopn = len(path) - 1
        arrow = "  ->  ".join(nodes[p]["name"] for p in path)
        direction = "BEN" if val > 0 else "LOSE"
        print(f"{i:<5}{val:>+8.2f}  {direction:<4}{hopn:>5}  {nodes[nid]['name'][:24]:<26}{arrow}  [{CH_LABEL.get(ch,ch)}]")
        rows.append({"rank": i, "node_id": nid, "name": nodes[nid]["name"],
                     "layer": nodes[nid]["layer_name"], "impact": round(val, 3),
                     "direction": "beneficiary" if val > 0 else "loser",
                     "hops": hopn, "path": " -> ".join(nodes[p]["name"] for p in path),
                     "channel": CH_LABEL.get(ch, ch)})
    if out_csv:
        with open(out_csv, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["rank", "node_id", "name", "layer", "impact",
                                              "direction", "hops", "path", "channel"])
            w.writeheader(); w.writerows(rows)
        print(f"\n  -> wrote {out_csv}")
    return rows

def main():
    ap = argparse.ArgumentParser(description="Catalyst / chain-effect analysis over the AI supply-chain graph.")
    ap.add_argument("--seeds"); ap.add_argument("--polarity", choices=["pos", "neg"], default="pos")
    ap.add_argument("--kind", choices=["demand", "supply", "macro"], default="demand")
    ap.add_argument("--hops", type=int, default=3); ap.add_argument("--decay", type=float, default=0.5)
    ap.add_argument("--out")
    a = ap.parse_args()
    nodes, edges = load_graph()
    links = build_links(nodes, edges)

    if not a.seeds:
        print("No --seeds given; running three illustrative demo catalysts.\n"
              "(Try:  python3 catalyst_analysis.py --seeds NVDA --polarity pos --kind demand)")
        run(["MSFT"], +1, "demand", nodes, links, a.hops, a.decay,
            label="Hyperscaler announces a giant new AI training-cluster order")
        run(["TSM"], -1, "supply", nodes, links, a.hops, a.decay,
            label="TSMC advanced-packaging (CoWoS) capacity is disrupted")
        run(["T_POWER"], +1, "macro", nodes, links, a.hops, a.decay,
            label="AI datacenter power demand surges (macro theme)")
        return
    seeds = resolve(a.seeds, nodes)
    if not seeds:
        print("No valid seeds resolved.", file=sys.stderr); sys.exit(1)
    run(seeds, +1 if a.polarity == "pos" else -1, a.kind, nodes, links, a.hops, a.decay, a.out)

if __name__ == "__main__":
    main()
